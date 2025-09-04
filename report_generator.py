# report_generator.py
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data_processor import REPORT_DEFINITIONS 

def export_to_excel_multi_sheet_report(
    df_agrupado_por_prefijo, 
    output_filepath, 
    report_type, 
    report_title_prefix="REPORTE",
    inventory_cols=None # ***** ASEGÚRATE QUE ESTE PARÁMETRO ESTÉ AQUÍ *****
):
    if df_agrupado_por_prefijo is None or df_agrupado_por_prefijo.empty:
        return False, "No hay datos para exportar."

    try:
        with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
            # --- Título y Fecha Común ---
            current_month_year = datetime.now().strftime("%B %Y").upper()
            title_report = f"{report_title_prefix} {current_month_year}"
            fecha_creacion_str = f"Fecha de Creación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # --- Hoja 1: Resumen por Subcategoría ---
            sheet1_name = "ResumenSubcategorias"
            resumen_data = []
            for _, row_prefijo in df_agrupado_por_prefijo.iterrows():
                resumen_data.append({
                    'Subcategoría (Prefijo Código)': row_prefijo['Codigo_Prefijo'],
                    'Cantidad Total Filas': row_prefijo['Cantidad_Total_General'],
                    'Valor Total General': row_prefijo['Venta_Total_General']
                })
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name=sheet1_name, index=False, startrow=2)
            ws_resumen = writer.sheets[sheet1_name]

            ws_resumen.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df_resumen.shape[1] if not df_resumen.empty else 1)
            cell_title_s1 = ws_resumen.cell(row=1, column=1, value=title_report)
            cell_title_s1.font = Font(bold=True, size=16); cell_title_s1.alignment = Alignment(horizontal='center')
            ws_resumen.merge_cells(start_row=2, start_column=1, end_row=2, end_column=df_resumen.shape[1] if not df_resumen.empty else 1)
            cell_date_s1 = ws_resumen.cell(row=2, column=1, value=fecha_creacion_str)
            cell_date_s1.font = Font(italic=True, size=10); cell_date_s1.alignment = Alignment(horizontal='center')
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            if not df_resumen.empty:
                for col_idx, value in enumerate(df_resumen.columns.values, 1):
                    cell = ws_resumen.cell(row=3, column=col_idx)
                    cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center', vertical='center')
                    column_letter = get_column_letter(col_idx)
                    max_len = max(df_resumen.iloc[:,col_idx-1].astype(str).map(len).max(), len(value)) + 3
                    ws_resumen.column_dimensions[column_letter].width = max_len

            # --- Hoja 2: Detalle Completo ---
            sheet2_name = "DetalleCompleto"
            ws_detalle = writer.book.create_sheet(title=sheet2_name)
            
            report_spec = REPORT_DEFINITIONS.get(report_type, {})
            
            detalle_cols_data_names = []    
            detalle_cols_display_headers = [] 
            
            if report_type == "Ventas":
                detalle_cols_data_names = report_spec.get("final_df_columns", [])
                if 'Unidad' in detalle_cols_data_names: detalle_cols_data_names.remove('Unidad')
                ordered_sales_internal_cols = ["Fecha de emisión", "Bodega", "Codigo", "Categoría Producto", "Cod. Catalogo", "Producto", "Cantidad", "Costo Venta", "Descripción", "% Descuento", "Total"]
                detalle_cols_data_names = [col for col in ordered_sales_internal_cols if col in detalle_cols_data_names]
                ventas_display_map = {"Cod. Catalogo": "Talla", "Producto": "Color"}
                detalle_cols_display_headers = [ventas_display_map.get(col, col) for col in detalle_cols_data_names]
            elif report_type == "Inventario":
                # Usar inventory_cols si se pasa, sino las de REPORT_DEFINITIONS
                detalle_cols_data_names = inventory_cols if inventory_cols else report_spec.get("gui_detail_display_columns_ordered", [])
                inv_display_map = report_spec.get("filter_column_display_map", {}) 
                detalle_cols_display_headers = [inv_display_map.get(col, col) for col in detalle_cols_data_names]
            else: 
                if not df_agrupado_por_prefijo.empty and df_agrupado_por_prefijo.iloc[0]['Detalles_Filas']:
                    detalle_cols_data_names = list(df_agrupado_por_prefijo.iloc[0]['Detalles_Filas'][0].keys())
                    if 'Codigo_Prefijo' in detalle_cols_data_names: detalle_cols_data_names.remove('Codigo_Prefijo')
                detalle_cols_display_headers = detalle_cols_data_names

            num_cols_detalle = len(detalle_cols_display_headers) if detalle_cols_display_headers else 1

            ws_detalle.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1,num_cols_detalle))
            cell_title_s2 = ws_detalle.cell(row=1, column=1, value=title_report); cell_title_s2.font = Font(bold=True, size=16); cell_title_s2.alignment = Alignment(horizontal='center')
            ws_detalle.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1,num_cols_detalle))
            cell_date_s2 = ws_detalle.cell(row=2, column=1, value=fecha_creacion_str); cell_date_s2.font = Font(italic=True, size=10); cell_date_s2.alignment = Alignment(horizontal='center')

            font_prefijo_header_val = Font(bold=True, size=11); font_prefijo_header_desc = Font(italic=True, size=11); fill_prefijo_header = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            font_detalle_table_header = Font(bold=True, color="FFFFFF"); fill_detalle_table_header = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            current_row_detalle = 4 

            for idx_prefijo_loop, (_, row_prefijo) in enumerate(df_agrupado_por_prefijo.iterrows()):
                fila_inicio_seccion_actual_detalle = current_row_detalle
                prefijo_info = [("Subcategoría (Prefijo Código):", row_prefijo['Codigo_Prefijo']), ("Cantidad Total Filas:", row_prefijo['Cantidad_Total_General']), (f"Valor Total (Suma Col. '{row_prefijo.get('total_col_name','Total')}'):", f"{row_prefijo['Venta_Total_General']:.2f}")]
                for desc, val in prefijo_info:
                    cell_desc = ws_detalle.cell(row=current_row_detalle, column=1, value=desc); cell_desc.font = font_prefijo_header_desc; cell_desc.fill = fill_prefijo_header; cell_desc.alignment = Alignment(horizontal='left')
                    cell_val = ws_detalle.cell(row=current_row_detalle, column=2, value=val); cell_val.font = font_prefijo_header_val; cell_val.fill = fill_prefijo_header
                    if num_cols_detalle > 2: ws_detalle.merge_cells(start_row=current_row_detalle, start_column=2, end_row=current_row_detalle, end_column=num_cols_detalle)
                    current_row_detalle += 1
                for c_idx in range(1, num_cols_detalle + 1): ws_detalle.cell(row=current_row_detalle, column=c_idx).border = Border(bottom=Side(style='thin'))
                current_row_detalle +=1

                link_cell_s1 = ws_resumen.cell(row=idx_prefijo_loop + 4, column=1)
                link_cell_s1.hyperlink = f"#'{sheet2_name}'!A{fila_inicio_seccion_actual_detalle}"; link_cell_s1.font = Font(color="0000FF", underline="single")

                if row_prefijo['Detalles_Filas'] and detalle_cols_display_headers and detalle_cols_display_headers[0] != 'Informacion_No_Disponible':
                    for col_idx, col_display_name in enumerate(detalle_cols_display_headers, 1): # Usar nombres de display
                        cell = ws_detalle.cell(row=current_row_detalle, column=col_idx, value=col_display_name)
                        cell.font = font_detalle_table_header; cell.fill = fill_detalle_table_header; cell.alignment = Alignment(horizontal='center', vertical='center')
                    current_row_detalle += 1
                    for detalle_fila_dict in row_prefijo['Detalles_Filas']:
                        for col_idx, col_data_name in enumerate(detalle_cols_data_names, 1): # Usar nombres internos para obtener datos
                            cell = ws_detalle.cell(row=current_row_detalle, column=col_idx)
                            cell_value = detalle_fila_dict.get(col_data_name, '')
                            if isinstance(cell_value, (int, float)):
                                if report_type == "Inventario" and col_data_name in ["Stock"]: cell.number_format = '0'
                                elif report_type == "Ventas" and col_data_name in ["Cantidad"]: cell.number_format = '0'
                                else: cell.number_format = '#,##0.00'
                            cell.value = cell_value
                        current_row_detalle += 1
                current_row_detalle += 1 

            if not df_agrupado_por_prefijo.empty and detalle_cols_display_headers and detalle_cols_display_headers[0] != 'Informacion_No_Disponible':
                for i, col_display_name in enumerate(detalle_cols_display_headers, 1):
                    max_len_data = 0; internal_col_name_for_width = detalle_cols_data_names[i-1] if i-1 < len(detalle_cols_data_names) else col_display_name
                    if row_prefijo['Detalles_Filas']:
                        for _, r_pref in df_agrupado_por_prefijo.iterrows():
                            for detalle_fila in r_pref['Detalles_Filas']: max_len_data = max(max_len_data, len(str(detalle_fila.get(internal_col_name_for_width, ''))))
                    column_letter = get_column_letter(i); max_len = max(max_len_data, len(col_display_name)) + 3
                    ws_detalle.column_dimensions[column_letter].width = max_len
        
        return True, f"Reporte multi-hoja ({report_type}) exportado a {output_filepath}"
    except Exception as e:
        import traceback
        print(f"Error detallado en export_to_excel_multi_sheet_report ({report_type}):")
        print(traceback.format_exc())
        return False, f"Error al exportar Excel ({report_type}): {e}"

def generate_bar_chart(df_agrupado_por_prefijo, output_image_path):
    if df_agrupado_por_prefijo is None or df_agrupado_por_prefijo.empty: return False, "No hay datos para generar el gráfico."
    try:
        if 'Codigo_Prefijo' not in df_agrupado_por_prefijo.columns or 'Venta_Total_General' not in df_agrupado_por_prefijo.columns: return False, "Columnas necesarias para gráfico no encontradas."
        df_for_chart = df_agrupado_por_prefijo 
        if df_for_chart.empty: return False, "No hay datos para el gráfico."
        fig, ax = plt.subplots(figsize=(max(10, len(df_for_chart)*0.5), 7)); prefijos = df_for_chart['Codigo_Prefijo'].astype(str)
        valores_totales = df_for_chart['Venta_Total_General']
        bars = ax.bar(prefijos, valores_totales, color='cornflowerblue', edgecolor='black')
        ax.set_xlabel("Subcategoría (Prefijo Código)", fontsize=12); ax.set_ylabel("Valor Total General", fontsize=12)
        ax.set_title("Valor Total General por Subcategoría (Prefijo Código)", fontsize=14, fontweight='bold')
        formatter = mticker.FuncFormatter(lambda x, p: f'{x:,.2f}'); ax.yaxis.set_major_formatter(formatter); ax.grid(axis='y', linestyle='--', alpha=0.7)
        for bar in bars: yval = bar.get_height(); 
        if abs(yval) > 0.001 : plt.text(bar.get_x() + bar.get_width()/2.0, yval + (valores_totales.max()*0.01) , f'{yval:,.0f}', ha='center', va='bottom', fontsize=9, color='dimgray')
        plt.xticks(rotation=45, ha="right", fontsize=10); plt.yticks(fontsize=10); plt.tight_layout(); plt.savefig(output_image_path, dpi=200, bbox_inches='tight'); plt.close(fig)
        return True, f"Gráfico exportado exitosamente a {output_image_path}"
    except Exception as e: import traceback; print(f"Error detallado en generate_bar_chart:"); print(traceback.format_exc()); return False, f"Error al generar el gráfico: {e}"